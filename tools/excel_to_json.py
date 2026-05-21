from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def convert_excel_to_json(source: Path, destination: Path, sheet_name: str | None = None) -> Path:
    workbook = load_workbook(source, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空")
    headers = [str(value).strip() for value in rows[0]]
    items = []
    for row in rows[1:]:
        if row is None or all(value is None for value in row):
            continue
        item = {header: value for header, value in zip(headers, row) if header}
        items.append(item)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    return destination


def main() -> None:
    parser = argparse.ArgumentParser(description="将 Excel 配置表转换为 JSON")
    parser.add_argument("source", type=Path, help="Excel 文件路径")
    parser.add_argument("destination", type=Path, help="JSON 输出路径")
    parser.add_argument("--sheet", dest="sheet_name", default=None, help="工作表名称")
    args = parser.parse_args()
    output = convert_excel_to_json(args.source, args.destination, args.sheet_name)
    print(f"已生成: {output}")


if __name__ == "__main__":
    main()

